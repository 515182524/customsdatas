#!/usr/bin/env python3
"""Generate a Markdown report and local SVG charts from China customs tables."""

import argparse
import csv
import datetime as dt
import html
import re
import statistics
from collections import defaultdict
from pathlib import Path


ALIASES = {
    "period": ["数据年月", "统计年月", "年月", "月份", "日期"],
    "amount": ["美元", "美元金额", "金额（美元）", "金额(美元)", "贸易额美元", "人民币", "金额"],
    "product_code": ["商品编码", "HS编码", "海关编码"],
    "product": ["商品名称", "商品"],
    "partner": ["贸易伙伴名称", "贸易伙伴", "产销国名称", "产销国", "国家地区"],
    "origin": ["原产国（地区）名称", "原产国(地区)名称", "原产国名称", "原产地"],
    "destination": ["最终目的国（地区）名称", "最终目的国(地区)名称", "最终目的国名称", "目的国名称", "目的国"],
    "region": ["注册地名称", "境内货源地", "境内目的地", "注册地"],
    "trade_mode": ["贸易方式名称", "贸易方式"],
    "flow": ["进出口", "进出口标志", "贸易流向"],
    "qty1": ["第一数量"],
    "unit1": ["第一计量单位"],
    "qty2": ["第二数量"],
    "unit2": ["第二计量单位"],
}

COLORS = ["#1769aa", "#2f8f9d", "#4f9d69", "#d89b31", "#c75c5c", "#6f63a8"]


def clean_header(value):
    return str(value or "").replace("\ufeff", "").strip()


def read_csv_file(path):
    last_error = None
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                rows = []
                for raw in reader:
                    row = {clean_header(k): v for k, v in raw.items() if clean_header(k)}
                    rows.append(row)
                if reader.fieldnames:
                    return rows, encoding
        except (UnicodeError, csv.Error) as exc:
            last_error = exc
    raise ValueError("无法识别 CSV 编码: {}".format(last_error))


def read_xlsx_file(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要 openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = [clean_header(v) for v in next(values)]
    rows = []
    for values_row in values:
        rows.append({headers[i]: value for i, value in enumerate(values_row) if i < len(headers) and headers[i]})
    return rows, "xlsx"


def read_table(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_file(path)
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx_file(path)
    raise ValueError("仅支持 CSV、XLSX、XLSM 文件")


def resolve_columns(headers):
    resolved = {}
    normalized = {clean_header(header): header for header in headers}
    for role, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                resolved[role] = normalized[alias]
                break
    return resolved


def number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("，", "")
    if not text or text in ("-", "--", "?", "N/A", "nan"):
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("%", "")
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return None


def normalize_period(value):
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.strftime("%Y-%m")
    digits = re.sub(r"\D", "", str(value))
    if len(digits) >= 6:
        year, month = digits[:4], digits[4:6]
        if month.isdigit() and 1 <= int(month) <= 12:
            return "{}-{}".format(year, month)
    return str(value).strip()


def month_key(period):
    match = re.fullmatch(r"(\d{4})-(\d{2})", period)
    if match:
        return int(match.group(1)) * 12 + int(match.group(2))
    return period


def period_sort_key(period):
    key = month_key(period)
    return (0, key) if isinstance(key, int) else (1, str(key))


def dimension_value(row, column):
    value = row.get(column, "")
    text = str(value).strip() if value is not None else ""
    return text or "未注明"


def aggregate(rows, column, amount_col):
    result = defaultdict(float)
    for row in rows:
        amount = number(row.get(amount_col))
        if amount is not None:
            result[dimension_value(row, column)] += amount
    return dict(result)


def top_items(mapping, limit):
    return sorted(mapping.items(), key=lambda item: item[1], reverse=True)[:limit]


def compact(value):
    absolute = abs(value)
    if absolute >= 100000000:
        return "{:.2f}亿".format(value / 100000000)
    if absolute >= 10000:
        return "{:.2f}万".format(value / 10000)
    return "{:,.0f}".format(value)


def money(value, label):
    if "美元" in label or label == "美元":
        return "${:,.0f}".format(value)
    return "{:,.0f}".format(value)


def percent(value):
    return "{:.1%}".format(value) if value is not None else "—"


def price(value, amount_label, unit):
    prefix = "$" if "美元" in amount_label or amount_label == "美元" else ""
    return "{}{:,.2f}/{}".format(prefix, value, unit)


def quantity(value, unit):
    return "{:,.0f} {}".format(value, unit)


def escape(value):
    return html.escape(str(value))


def svg_start(width, height, title):
    return [
        '<svg xmlns="http://www.w3.org/2000/svg" width="{}" height="{}" viewBox="0 0 {} {}">'.format(width, height, width, height),
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        '<text x="{}" y="28" text-anchor="middle" font-family="Arial, PingFang SC, sans-serif" font-size="18" font-weight="600" fill="#1f2937">{}</text>'.format(width / 2, escape(title)),
    ]


def write_bar_svg(items, title, path, diverging=False, value_formatter=compact):
    items = list(items)
    width, left, right, top, row_h = 920, 180, 110, 52, 34
    height = max(150, top + row_h * len(items) + 35)
    lines = svg_start(width, height, title)
    values = [value for _, value in items] or [0]
    if diverging:
        max_abs = max(abs(value) for value in values) or 1
        center = left + (width - left - right) / 2
        half = (width - left - right) / 2
        lines.append('<line x1="{0}" y1="{1}" x2="{0}" y2="{2}" stroke="#9ca3af"/>'.format(center, top - 12, height - 25))
    else:
        max_abs = max(values) or 1
    for index, (label, value) in enumerate(items):
        y = top + index * row_h
        lines.append('<text x="{}" y="{}" text-anchor="end" font-family="Arial, PingFang SC, sans-serif" font-size="13" fill="#374151">{}</text>'.format(left - 10, y + 17, escape(label[:18])))
        if diverging:
            bar_w = abs(value) / max_abs * half
            x = center if value >= 0 else center - bar_w
            color = "#2f8f9d" if value >= 0 else "#c75c5c"
            text_x = center + bar_w + 6 if value >= 0 else center - bar_w - 6
            anchor = "start" if value >= 0 else "end"
        else:
            bar_w = max(1, value / max_abs * (width - left - right))
            x, color, text_x, anchor = left, COLORS[index % len(COLORS)], left + bar_w + 6, "start"
        lines.append('<rect x="{}" y="{}" width="{}" height="22" rx="3" fill="{}"/>'.format(x, y, bar_w, color))
        lines.append('<text x="{}" y="{}" text-anchor="{}" font-family="Arial, PingFang SC, sans-serif" font-size="12" fill="#4b5563">{}</text>'.format(text_x, y + 16, anchor, escape(value_formatter(value))))
    lines.append("</svg>")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_line_svg(items, title, path, value_formatter=compact):
    items = list(items)
    width, height, left, right, top, bottom = 920, 420, 80, 35, 55, 70
    plot_w, plot_h = width - left - right, height - top - bottom
    values = [value for _, value in items]
    maximum = max(values) or 1
    lines = svg_start(width, height, title)
    for tick in range(5):
        y = top + plot_h * tick / 4
        tick_value = maximum * (4 - tick) / 4
        lines.append('<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="#e5e7eb"/>'.format(left, y, width - right, y))
        lines.append('<text x="{}" y="{}" text-anchor="end" font-family="Arial, sans-serif" font-size="11" fill="#6b7280">{}</text>'.format(left - 8, y + 4, escape(value_formatter(tick_value))))
    points = []
    count = max(1, len(items) - 1)
    for index, (label, value) in enumerate(items):
        x = left + plot_w * index / count
        y = top + plot_h * (1 - value / maximum)
        points.append((x, y))
        lines.append('<text x="{}" y="{}" text-anchor="middle" font-family="Arial, sans-serif" font-size="11" fill="#6b7280">{}</text>'.format(x, height - 38, escape(label)))
        lines.append('<circle cx="{}" cy="{}" r="4" fill="#1769aa"/>'.format(x, y))
        lines.append('<text x="{}" y="{}" text-anchor="middle" font-family="Arial, sans-serif" font-size="11" fill="#374151">{}</text>'.format(x, y - 10, escape(value_formatter(value))))
    lines.append('<polyline fill="none" stroke="#1769aa" stroke-width="3" points="{}"/>'.format(" ".join("{},{}".format(x, y) for x, y in points)))
    lines.append("</svg>")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_heatmap_svg(periods, groups, matrix, title, path):
    cell_w, cell_h, left, top = 78, 32, 180, 62
    width = left + cell_w * len(periods) + 45
    height = top + cell_h * len(groups) + 45
    maximum = max(matrix.values()) if matrix else 1
    lines = svg_start(width, height, title)
    for col, period in enumerate(periods):
        lines.append('<text x="{}" y="{}" text-anchor="middle" font-family="Arial, sans-serif" font-size="11" fill="#4b5563">{}</text>'.format(left + col * cell_w + cell_w / 2, top - 10, escape(period)))
    for row_index, group in enumerate(groups):
        y = top + row_index * cell_h
        lines.append('<text x="{}" y="{}" text-anchor="end" font-family="Arial, PingFang SC, sans-serif" font-size="12" fill="#374151">{}</text>'.format(left - 10, y + 21, escape(group[:18])))
        for col, period in enumerate(periods):
            value = matrix.get((group, period), 0)
            intensity = value / maximum if maximum else 0
            blue = int(245 - intensity * 145)
            color = "rgb({},{},{})".format(225 - int(intensity * 110), 238 - int(intensity * 90), blue)
            x = left + col * cell_w
            lines.append('<rect x="{}" y="{}" width="{}" height="{}" fill="{}" stroke="#ffffff"/>'.format(x, y, cell_w, cell_h, color))
            if value:
                lines.append('<text x="{}" y="{}" text-anchor="middle" font-family="Arial, sans-serif" font-size="10" fill="#1f2937">{}</text>'.format(x + cell_w / 2, y + 20, escape(compact(value))))
    lines.append("</svg>")
    path.write_text("\n".join(lines), encoding="utf-8")


def markdown_table(headers, rows):
    output = ["| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
    for row in rows:
        output.append("| " + " | ".join(str(value).replace("|", "\\|") for value in row) + " |")
    return "\n".join(output)


def previous_period(period):
    key = month_key(period)
    if not isinstance(key, int):
        return None
    key -= 1
    return "{:04d}-{:02d}".format((key - 1) // 12, (key - 1) % 12 + 1)


def last_year_period(period):
    match = re.fullmatch(r"(\d{4})-(\d{2})", period)
    return "{}-{}".format(int(match.group(1)) - 1, match.group(2)) if match else None


def quantity_specs(columns, rows):
    specs = []
    for index in (1, 2):
        qty_role, unit_role = "qty{}".format(index), "unit{}".format(index)
        if qty_role not in columns or unit_role not in columns:
            continue
        units = sorted(set(
            dimension_value(row, columns[unit_role])
            for row in rows
            if number(row.get(columns[qty_role])) is not None and number(row.get(columns[qty_role])) > 0
        ))
        specs.append({
            "index": index,
            "qty_col": columns[qty_role],
            "unit_col": columns[unit_role],
            "units": units,
            "unit": units[0] if len(units) == 1 else "",
        })
    return specs


def quantity_price_stats(rows, amount_col, qty_col):
    valid = []
    for row in rows:
        amount_value = number(row.get(amount_col))
        qty_value = number(row.get(qty_col))
        if amount_value is not None and qty_value is not None and qty_value > 0:
            valid.append((row, amount_value, qty_value))
    amount_total = sum(item[1] for item in valid)
    qty_total = sum(item[2] for item in valid)
    return {
        "valid": valid,
        "amount": amount_total,
        "quantity": qty_total,
        "unit_price": amount_total / qty_total if qty_total else None,
    }


def quantity_price_by_dimension(valid, column):
    grouped = defaultdict(lambda: [0.0, 0.0])
    for row, amount_value, qty_value in valid:
        key = dimension_value(row, column)
        grouped[key][0] += amount_value
        grouped[key][1] += qty_value
    return {
        key: {"amount": values[0], "quantity": values[1], "unit_price": values[0] / values[1]}
        for key, values in grouped.items()
        if values[1] > 0
    }


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="CSV/XLSX path")
    parser.add_argument("--output", default=None, help="Output directory")
    parser.add_argument("--title", default="中国海关贸易数据多维分析报告")
    parser.add_argument("--flow-label", default="", help="Confirmed flow label, such as 出口 or 进口")
    parser.add_argument("--top-n", type=int, default=10)
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output = Path(args.output).expanduser().resolve() if args.output else Path.cwd() / "{}-report".format(input_path.stem)
    chart_dir = output / "charts"
    chart_dir.mkdir(parents=True, exist_ok=True)

    rows, source_format = read_table(input_path)
    if not rows:
        raise ValueError("输入表格没有数据行")
    columns = resolve_columns(rows[0].keys())
    if "product" not in columns and "product_code" in columns:
        columns["product"] = columns["product_code"]
    if "amount" not in columns:
        raise ValueError("未识别到金额列；请检查 references/customs-fields.md 并补充字段别名")
    amount_col = columns["amount"]

    valid_rows = [row for row in rows if number(row.get(amount_col)) is not None]
    total = sum(number(row.get(amount_col)) or 0 for row in valid_rows)
    report = ["# {}".format(args.title), ""]
    report.append("> 数据源：`{}`；生成时间：{}；有效金额记录：{:,}/{:,}。".format(input_path.name, dt.datetime.now().strftime("%Y-%m-%d %H:%M"), len(valid_rows), len(rows)))
    report.extend(["", "## 口径与数据质量", ""])
    report.append("- 文件格式/编码：`{}`；金额字段：`{}`。".format(source_format, amount_col))

    periods = []
    period_totals = {}
    if "period" in columns:
        period_totals_raw = defaultdict(float)
        for row in valid_rows:
            period_totals_raw[normalize_period(row.get(columns["period"]))] += number(row.get(amount_col)) or 0
        period_totals = dict(period_totals_raw)
        periods = sorted(period_totals, key=period_sort_key)
        report.append("- 时间范围：{} 至 {}，共 {} 个时期。".format(periods[0], periods[-1], len(periods)))
    else:
        report.append("- 未识别到时间字段，因此不生成趋势与变化拆解。")

    if args.flow_label:
        report.append("- 用户确认的贸易流向：**{}**。".format(args.flow_label))
    elif "flow" in columns:
        flows = sorted(set(dimension_value(row, columns["flow"]) for row in rows))
        report.append("- 表内贸易流向字段：`{}`；取值：{}。".format(columns["flow"], "、".join(flows[:8])))
    else:
        report.append("- 表内没有明确进出口方向字段，无法仅凭此表判断进口或出口。")

    if "origin" not in columns and "destination" not in columns and "partner" in columns:
        report.append("- 仅识别到贸易伙伴字段；贸易伙伴不自动等同于原产国或最终目的国。")
    missing_amount = len(rows) - len(valid_rows)
    if missing_amount:
        report.append("- 有 {:,} 行金额为空或无法解析，已排除。".format(missing_amount))
    if len(periods) >= 3:
        previous_values = [period_totals[p] for p in periods[-4:-1]]
        baseline = sorted(previous_values)[len(previous_values) // 2] if previous_values else 0
        if baseline and period_totals[periods[-1]] < baseline * 0.7:
            report.append("- **提示：最新时期金额明显低于此前中位水平，可能是未完结月份，需谨慎解释下降。**")

    dimension_maps = {}
    for role in ("partner", "origin", "destination", "product", "region", "trade_mode"):
        if role in columns:
            dimension_maps[role] = aggregate(valid_rows, columns[role], amount_col)
    qty_specs = quantity_specs(columns, valid_rows)
    qty_analyses = []
    for spec in qty_specs:
        if len(spec["units"]) == 1:
            analysis = dict(spec)
            analysis.update(quantity_price_stats(valid_rows, amount_col, spec["qty_col"]))
            qty_analyses.append(analysis)
        elif not spec["units"]:
            report.append("- `{}` 没有大于 0 的有效数量，不计算汇总数量与单价。".format(spec["qty_col"]))
        else:
            report.append("- `{}` 包含多个计量单位（{}），不计算汇总数量与单价。".format(spec["qty_col"], "、".join(spec["units"])))

    report.extend(["", "## 核心摘要", ""])
    summary = [["总金额", money(total, amount_col)], ["有效记录", "{:,}".format(len(valid_rows))]]
    if periods:
        summary.append(["时间范围", "{} 至 {}".format(periods[0], periods[-1])])
        summary.append(["最新时期金额", money(period_totals[periods[-1]], amount_col)])
    partner_role = "partner" if "partner" in dimension_maps else ("destination" if "destination" in dimension_maps else ("origin" if "origin" in dimension_maps else None))
    if partner_role:
        ranked = top_items(dimension_maps[partner_role], 5)
        summary.append(["首位国家/地区", "{}（{}）".format(ranked[0][0], percent(ranked[0][1] / total if total else 0))])
        summary.append(["CR5", percent(sum(value for _, value in ranked) / total if total else 0)])
    if "region" in dimension_maps:
        item = top_items(dimension_maps["region"], 1)[0]
        summary.append(["首位国内区域", "{}（{}）".format(item[0], percent(item[1] / total if total else 0))])
    for analysis in qty_analyses:
        summary.append([
            "加权单价（{}）".format(analysis["unit"]),
            price(analysis["unit_price"], amount_col, analysis["unit"]),
        ])
    report.append(markdown_table(["指标", "结果"], summary))

    report.extend(["", "## 月度趋势", ""])
    if len(periods) >= 2:
        trend_items = [(period, period_totals[period]) for period in periods]
        write_line_svg(trend_items, "月度金额趋势", chart_dir / "monthly-trend.svg")
        report.append("![月度金额趋势](charts/monthly-trend.svg)")
        trend_rows = []
        for period in periods:
            prior = previous_period(period)
            last_year = last_year_period(period)
            mom = period_totals[period] / period_totals[prior] - 1 if prior in period_totals and period_totals[prior] else None
            yoy = period_totals[period] / period_totals[last_year] - 1 if last_year in period_totals and period_totals[last_year] else None
            trend_rows.append([period, money(period_totals[period], amount_col), percent(mom), percent(yoy)])
        report.append(markdown_table(["时期", "金额", "环比", "同比"], trend_rows))
    elif periods:
        report.append("仅有一个时期 `{}`，不生成趋势图。该时期金额为 {}。".format(periods[0], money(period_totals[periods[0]], amount_col)))
    else:
        report.append("缺少可用时间字段。")

    labels = {
        "partner": "贸易伙伴",
        "origin": "原产地",
        "destination": "目的国",
        "product": "商品",
        "region": columns.get("region", "国内区域"),
        "trade_mode": "贸易方式",
    }
    for role in ("partner", "origin", "destination", "product", "region", "trade_mode"):
        if role not in dimension_maps:
            continue
        mapping = dimension_maps[role]
        label = labels[role]
        report.extend(["", "## {}分析".format(label), ""])
        ranked = top_items(mapping, args.top_n)
        if len(mapping) > 1:
            filename = "{}-top.svg".format(role)
            write_bar_svg(ranked, "{}金额 Top {}".format(label, min(args.top_n, len(ranked))), chart_dir / filename)
            report.append("![{}金额排名](charts/{})".format(label, filename))
        else:
            report.append("该维度只有一个取值，不生成排名图。")
        table_rows = [[index + 1, name, money(value, amount_col), percent(value / total if total else 0)] for index, (name, value) in enumerate(ranked)]
        report.append(markdown_table(["排名", label, "金额", "份额"], table_rows))
        shares = [(value / total) for value in mapping.values()] if total else []
        hhi = sum(share * share for share in shares) * 10000
        report.append("{}共 **{}** 个取值，CR3 为 **{}**，CR5 为 **{}**，HHI 为 **{:.0f}**。".format(
            label,
            len(mapping),
            percent(sum(value for _, value in top_items(mapping, 3)) / total if total else 0),
            percent(sum(value for _, value in top_items(mapping, 5)) / total if total else 0),
            hhi,
        ))

    if partner_role and len(periods) >= 2:
        role_col = columns[partner_role]
        group_period = defaultdict(float)
        for row in valid_rows:
            group = dimension_value(row, role_col)
            period = normalize_period(row.get(columns["period"]))
            group_period[(group, period)] += number(row.get(amount_col)) or 0
        top_groups = [name for name, _ in top_items(dimension_maps[partner_role], min(8, args.top_n))]
        write_heatmap_svg(periods, top_groups, group_period, "{} × 月份金额热力图".format(labels[partner_role]), chart_dir / "{}-monthly-heatmap.svg".format(partner_role))
        report.extend(["", "## 时间与市场交叉分析", "", "![时间与市场热力图](charts/{}-monthly-heatmap.svg)".format(partner_role)])

        first, last = periods[0], periods[-1]
        changes = []
        for group in dimension_maps[partner_role]:
            change = group_period.get((group, last), 0) - group_period.get((group, first), 0)
            if change:
                changes.append((group, change))
        changes = sorted(changes, key=lambda item: abs(item[1]), reverse=True)[:args.top_n]
        if changes:
            write_bar_svg(changes, "{}至{}金额变化贡献".format(first, last), chart_dir / "{}-change.svg".format(partner_role), diverging=True)
            report.extend(["", "## 首末期变化驱动", "", "![首末期金额变化贡献](charts/{}-change.svg)".format(partner_role)])
            report.append(markdown_table([labels[partner_role], "金额变化"], [[name, money(value, amount_col)] for name, value in changes]))

    cross_pairs = [("partner", "product"), ("partner", "region"), ("origin", "product"), ("destination", "product")]
    cross_role = next(((a, b) for a, b in cross_pairs if a in columns and b in columns and len(dimension_maps.get(a, {})) > 1 and len(dimension_maps.get(b, {})) > 1), None)
    if cross_role:
        first_role, second_role = cross_role
        cross = defaultdict(float)
        for row in valid_rows:
            key = (dimension_value(row, columns[first_role]), dimension_value(row, columns[second_role]))
            cross[key] += number(row.get(amount_col)) or 0
        report.extend(["", "## 主要交叉组合", ""])
        cross_rows = [[a, b, money(value, amount_col), percent(value / total if total else 0)] for (a, b), value in top_items(cross, args.top_n)]
        report.append(markdown_table([labels[first_role], labels[second_role], "金额", "份额"], cross_rows))

    for analysis in qty_analyses:
        unit = analysis["unit"]
        unit_slug = "qty{}".format(analysis["index"])
        coverage = analysis["amount"] / total if total else 0
        report.extend(["", "## 数量与价格分析：{}".format(unit), ""])
        report.append(markdown_table(
            ["指标", "结果"],
            [
                ["对应字段", "{} / {}".format(analysis["qty_col"], analysis["unit_col"])],
                ["有效数量记录", "{:,}/{:,}".format(len(analysis["valid"]), len(valid_rows))],
                ["有效金额覆盖率", percent(coverage)],
                ["数量合计", quantity(analysis["quantity"], unit)],
                ["加权单价", price(analysis["unit_price"], amount_col, unit)],
            ],
        ))
        report.append("> 加权单价 = 该数量大于 0 的记录金额合计 / 数量合计；不同市场、规格和贸易方式的结构变化也会影响该指标。")

        if len(periods) >= 2:
            monthly = defaultdict(lambda: [0.0, 0.0])
            for row, amount_value, qty_value in analysis["valid"]:
                period = normalize_period(row.get(columns["period"]))
                monthly[period][0] += amount_value
                monthly[period][1] += qty_value
            monthly_prices = {period: values[0] / values[1] for period, values in monthly.items() if values[1] > 0}
            monthly_qty = {period: values[1] for period, values in monthly.items()}
            chart_periods = [period for period in periods if period in monthly_prices]
            write_line_svg(
                [(period, monthly_qty[period]) for period in chart_periods],
                "月度{}数量趋势".format(unit),
                chart_dir / "{}-quantity-trend.svg".format(unit_slug),
            )
            write_line_svg(
                [(period, monthly_prices[period]) for period in chart_periods],
                "月度加权单价趋势（{}/{}）".format(amount_col, unit),
                chart_dir / "{}-unit-price-trend.svg".format(unit_slug),
                value_formatter=lambda value: "{:,.2f}".format(value),
            )
            report.append("![月度{}数量趋势](charts/{}-quantity-trend.svg)".format(unit, unit_slug))
            report.append("![月度{}加权单价趋势](charts/{}-unit-price-trend.svg)".format(unit, unit_slug))

            linkage_rows = []
            for period in chart_periods:
                prior = previous_period(period)
                qty_mom = monthly_qty[period] / monthly_qty[prior] - 1 if prior in monthly_qty and monthly_qty[prior] else None
                price_mom = monthly_prices[period] / monthly_prices[prior] - 1 if prior in monthly_prices and monthly_prices[prior] else None
                linkage_rows.append([
                    period,
                    money(monthly[period][0], amount_col),
                    quantity(monthly_qty[period], unit),
                    price(monthly_prices[period], amount_col, unit),
                    percent(qty_mom),
                    percent(price_mom),
                ])
            report.append(markdown_table(
                ["时期", "金额", "数量", "加权单价", "数量环比", "单价环比"],
                linkage_rows,
            ))

            price_values = [monthly_prices[period] for period in chart_periods]
            mean_price = statistics.mean(price_values)
            volatility = statistics.pstdev(price_values) / mean_price if mean_price and len(price_values) > 1 else 0
            min_period = min(chart_periods, key=lambda period: monthly_prices[period])
            max_period = max(chart_periods, key=lambda period: monthly_prices[period])
            first_period, last_period = chart_periods[0], chart_periods[-1]
            first_last_change = monthly_prices[last_period] / monthly_prices[first_period] - 1 if monthly_prices[first_period] else None
            report.extend(["", "### {}价格波动摘要".format(unit), ""])
            report.append(markdown_table(
                ["指标", "结果"],
                [
                    ["月度单价最低", "{}：{}".format(min_period, price(monthly_prices[min_period], amount_col, unit))],
                    ["月度单价最高", "{}：{}".format(max_period, price(monthly_prices[max_period], amount_col, unit))],
                    ["首末期单价变化", percent(first_last_change)],
                    ["月度单价变异系数", percent(volatility)],
                ],
            ))

        if partner_role:
            partner_prices = quantity_price_by_dimension(analysis["valid"], columns[partner_role])
            top_partner_names = [name for name, _ in top_items(dimension_maps[partner_role], args.top_n) if name in partner_prices]
            partner_rows = []
            partner_chart = []
            for name in top_partner_names:
                stats = partner_prices[name]
                partner_rows.append([
                    name,
                    money(stats["amount"], amount_col),
                    quantity(stats["quantity"], unit),
                    price(stats["unit_price"], amount_col, unit),
                ])
                partner_chart.append((name, stats["unit_price"]))
            if partner_rows:
                write_bar_svg(
                    partner_chart,
                    "主要{}{}加权单价".format(labels[partner_role], unit),
                    chart_dir / "{}-partner-unit-price.svg".format(unit_slug),
                    value_formatter=lambda value: "{:,.2f}".format(value),
                )
                report.extend(["", "### 主要市场{}单价对比".format(unit), ""])
                report.append("![主要市场{}单价对比](charts/{}-partner-unit-price.svg)".format(unit, unit_slug))
                report.append(markdown_table(
                    [labels[partner_role], "有效金额", "数量", "加权单价"],
                    partner_rows,
                ))

    if len(qty_analyses) >= 2:
        first_qty, second_qty = qty_analyses[0], qty_analyses[1]
        paired = []
        for row in valid_rows:
            first_value = number(row.get(first_qty["qty_col"]))
            second_value = number(row.get(second_qty["qty_col"]))
            if first_value is not None and second_value is not None and first_value > 0 and second_value > 0:
                paired.append((row, first_value, second_value))
        first_total = sum(item[1] for item in paired)
        second_total = sum(item[2] for item in paired)
        if first_total > 0:
            ratio_label = "{}/{}".format(second_qty["unit"], first_qty["unit"])
            overall_ratio = second_total / first_total
            report.extend(["", "## 双数量关联分析", ""])
            report.append(markdown_table(
                ["指标", "结果"],
                [
                    ["双数量有效记录", "{:,}/{:,}".format(len(paired), len(valid_rows))],
                    ["{}合计".format(first_qty["unit"]), quantity(first_total, first_qty["unit"])],
                    ["{}合计".format(second_qty["unit"]), quantity(second_total, second_qty["unit"])],
                    ["平均{}".format(ratio_label), "{:,.2f} {}".format(overall_ratio, ratio_label)],
                ],
            ))
            report.append("> 双数量比值用于描述平均规格或重量关系；其变化可能来自产品组合变化，不代表单个产品规格变化。")
            if len(periods) >= 2:
                paired_monthly = defaultdict(lambda: [0.0, 0.0])
                for row, first_value, second_value in paired:
                    period = normalize_period(row.get(columns["period"]))
                    paired_monthly[period][0] += first_value
                    paired_monthly[period][1] += second_value
                ratio_periods = [period for period in periods if paired_monthly[period][0] > 0]
                monthly_ratios = {
                    period: paired_monthly[period][1] / paired_monthly[period][0]
                    for period in ratio_periods
                }
                write_line_svg(
                    [(period, monthly_ratios[period]) for period in ratio_periods],
                    "月度平均{}".format(ratio_label),
                    chart_dir / "dual-quantity-ratio-trend.svg",
                    value_formatter=lambda value: "{:,.2f}".format(value),
                )
                report.append("![双数量比值趋势](charts/dual-quantity-ratio-trend.svg)")
                report.append(markdown_table(
                    ["时期", "{}数量".format(first_qty["unit"]), "{}数量".format(second_qty["unit"]), "平均{}".format(ratio_label)],
                    [
                        [
                            period,
                            quantity(paired_monthly[period][0], first_qty["unit"]),
                            quantity(paired_monthly[period][1], second_qty["unit"]),
                            "{:,.2f} {}".format(monthly_ratios[period], ratio_label),
                        ]
                        for period in ratio_periods
                    ],
                ))

    report.extend(["", "## 解读边界", ""])
    report.append("- 本报告基于表内聚合结果，只描述数据表现，不推断政策、需求、竞争或供应链变化的原因。")
    report.append("- 排名与集中度受当前查询的商品、时间、贸易流向和字段口径影响。")
    if not args.flow_label and "flow" not in columns:
        report.append("- 如需区分进口/出口，请补充查询口径或包含进出口方向的字段。")
    if "origin" not in columns or "destination" not in columns:
        report.append("- 原产地或最终目的国字段不完整，缺失维度未做推断。")

    report_path = output / "report.md"
    report_path.write_text("\n".join(report) + "\n", encoding="utf-8")
    print(report_path)


if __name__ == "__main__":
    main()
